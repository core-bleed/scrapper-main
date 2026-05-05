"""Export logic for leads (Week 8)."""
from __future__ import annotations

import csv
import io
import uuid

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.db import models as m

EXPORT_FIELDS = [
    "person_name",
    "first_name",
    "last_name",
    "linkedin_url",
    "company_name",
    "company_domain",
    "title",
    "seniority",
    "is_c_suite",
    "best_email",
    "email_status",
    "email_confidence",
    "pattern_code",
    "domain_confidence",
    "verified_at",
]


async def build_export_rows(
    session: AsyncSession,
    person_ids: list[uuid.UUID],
) -> list[dict]:
    if not person_ids:
        return []

    stmt = (
        select(m.Person)
        .where(m.Person.id.in_(person_ids))
        .options(
            selectinload(m.Person.employments).selectinload(m.Employment.company),
            selectinload(m.Person.contact_candidates),
        )
        .order_by(m.Person.full_name)
    )
    r = await session.execute(stmt)
    people = list(r.scalars().unique().all())

    rows: list[dict] = []
    for person in people:
        emps = sorted(
            person.employments,
            key=lambda e: (e.title_rank, int(e.is_current)),
            reverse=True,
        )
        emp = emps[0] if emps else None
        company = emp.company if emp else None

        best_cand = next(
            (c for c in person.contact_candidates if c.is_best_current),
            next(
                (c for c in person.contact_candidates if emp and c.id == emp.best_candidate_id),
                None,
            ),
        )

        verified_at = (
            best_cand.latest_verified_at.isoformat()
            if best_cand and best_cand.latest_verified_at
            else ""
        )

        rows.append({
            "person_name": person.full_name,
            "first_name": person.first_name or "",
            "last_name": person.last_name or "",
            "linkedin_url": person.linkedin_url or "",
            "company_name": company.name if company else "",
            "company_domain": company.primary_domain if company else "",
            "title": emp.title if emp else "",
            "seniority": emp.seniority if emp else "",
            "is_c_suite": "true" if emp and emp.is_c_suite else "false",
            "best_email": best_cand.email if best_cand else "",
            "email_status": best_cand.verification_state if best_cand else "",
            "email_confidence": "",
            "pattern_code": best_cand.pattern_code if best_cand else "",
            "domain_confidence": (
                f"{float(company.domain_confidence):.4f}"
                if company and company.domain_confidence
                else ""
            ),
            "verified_at": verified_at,
        })

    return rows


def rows_to_csv(rows: list[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=EXPORT_FIELDS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def rows_to_xlsx(rows: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "SDR Leads"
    bold = Font(bold=True)

    for col, field in enumerate(EXPORT_FIELDS, start=1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font = bold

    for ri, row in enumerate(rows, start=2):
        for col, field in enumerate(EXPORT_FIELDS, start=1):
            ws.cell(row=ri, column=col, value=row.get(field, ""))

    for col in range(1, len(EXPORT_FIELDS) + 1):
        letter = ws.cell(row=1, column=col).column_letter
        maxlen = len(EXPORT_FIELDS[col - 1])
        for cell_row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            v = cell_row[0].value
            if v:
                maxlen = max(maxlen, min(len(str(v)), 60))
        ws.column_dimensions[letter].width = maxlen + 2

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
