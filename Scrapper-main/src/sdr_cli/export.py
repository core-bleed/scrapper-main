import csv
from pathlib import Path

import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font

from sdr_cli.db import export_rows_query
from sdr_cli.models import ExportRow


EXPORT_FIELDS = [
    "company_name",
    "company_domain",
    "company_batch",
    "company_source",
    "person_name",
    "title",
    "seniority",
    "linkedin_url",
    "twitter_url",
    "work_email",
    "email_status",
]


def export_csv(
    conn: sqlite3.Connection,
    output_path: Path,
    has_linkedin: bool = False,
    has_email: bool = False,
    source: str | None = None,
) -> int:
    rows = export_rows_query(conn, has_linkedin, has_email, source)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=EXPORT_FIELDS, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r.model_dump())
    return len(rows)


def export_xlsx(
    conn: sqlite3.Connection,
    output_path: Path,
    has_linkedin: bool = False,
    has_email: bool = False,
    source: str | None = None,
) -> int:
    rows = export_rows_query(conn, has_linkedin, has_email, source)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    bold = Font(bold=True)
    for col, h in enumerate(EXPORT_FIELDS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
    for ri, row in enumerate(rows, start=2):
        data = row.model_dump()
        for col, h in enumerate(EXPORT_FIELDS, start=1):
            ws.cell(row=ri, column=col, value=data.get(h))
    for col in range(1, len(EXPORT_FIELDS) + 1):
        letter = ws.cell(row=1, column=col).column_letter
        maxlen = len(EXPORT_FIELDS[col - 1])
        for r in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            v = r[0].value
            if v is not None:
                maxlen = max(maxlen, min(len(str(v)), 60))
        ws.column_dimensions[letter].width = maxlen + 2
    ws.freeze_panes = "A2"
    wb.save(output_path)
    return len(rows)


def rows_to_dicts(rows: list[ExportRow]) -> list[dict]:
    return [r.model_dump() for r in rows]
